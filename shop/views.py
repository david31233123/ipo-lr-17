from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import login
from django.db.models import Q
from django.core.paginator import Paginator
from .models import (
    Товар, Категория, Производитель, Корзина, ЭлементКорзины,
    Заказ, ЭлементЗаказа, Profile
)
from django.contrib.auth.models import User
from django.core.mail import send_mail
from django.conf import settings
from django.http import HttpResponse
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO

from rest_framework import permissions, status, viewsets
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from .serializers import (
    ProductSerializer,
    CategorySerializer,
    ManufacturerSerializer,
    CartSerializer,
    CartItemSerializer,
    ProfileSerializer,
    OrderSerializer,
)
from .forms import UserRegistrationForm


class IsAdminOrReadOnly(permissions.BasePermission):
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return request.user and request.user.is_authenticated and request.user.is_staff


# ==================== Страницы из предыдущего задания ====================

def home(request):
    """Главная страница магазина игрушек"""
    popular_products = Товар.objects.select_related('категория', 'производитель').order_by('-id')[:6]
    categories = Категория.objects.all()

    context = {
        'popular_products': popular_products,
        'categories': categories,
    }
    return render(request, 'shop/index.html', context)


def about_author(request):
    """Страница об авторе"""
    return render(request, 'shop/about.html')


def store_info(request):
    """Страница о магазине настольных игр"""
    return render(request, 'shop/store.html')


def register(request):
    """Регистрация нового пользователя"""
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            profile = user.profile
            profile.full_name = form.cleaned_data.get('full_name', '')
            profile.phone = form.cleaned_data.get('phone', '')
            profile.address = form.cleaned_data.get('address', '')
            profile.city = form.cleaned_data.get('city', '')
            profile.postal_code = form.cleaned_data.get('postal_code', '')
            profile.favorite_category = form.cleaned_data.get('favorite_category')
            profile.save()
            login(request, user)
            messages.success(request, 'Регистрация прошла успешно. Добро пожаловать!')
            return redirect('home')
    else:
        form = UserRegistrationForm()

    return render(request, 'shop/register.html', {'form': form})


@login_required
def profile_view(request):
    return render(request, 'shop/profile.html')


@login_required
def settings_view(request):
    profile = request.user.profile
    if request.method == 'POST':
        profile.full_name = request.POST.get('full_name', profile.full_name)
        profile.phone = request.POST.get('phone', profile.phone)
        profile.address = request.POST.get('address', profile.address)
        profile.city = request.POST.get('city', profile.city)
        profile.postal_code = request.POST.get('postal_code', profile.postal_code)
        profile.save()
        messages.success(request, 'Профиль обновлен.')
        return redirect('profile')

    return render(request, 'shop/settings.html')


@login_required
def order_detail(request, order_id):
    заказ = get_object_or_404(Заказ, pk=order_id, пользователь=request.user)
    return render(request, 'shop/order_detail.html', {'заказ': заказ})


# ==================== Задание 2: Представления для каталога и корзины ====================

def product_list(request):
    """
    Каталог товаров с фильтрацией и пагинацией
    """
    товары = Товар.objects.select_related('категория', 'производитель').all()

    категория_id = request.GET.get('category')
    производитель_id = request.GET.get('manufacturer')
    поиск = request.GET.get('search')
    page_number = request.GET.get('page')

    if категория_id:
        товары = товары.filter(категория_id=категория_id)
    if производитель_id:
        товары = товары.filter(производитель_id=производитель_id)
    if поиск:
        товары = товары.filter(
            Q(название__icontains=поиск) | Q(описание__icontains=поиск)
        )

    paginator = Paginator(товары, 9)
    page_obj = paginator.get_page(page_number)

    категории = Категория.objects.all()
    производители = Производитель.objects.all()

    context = {
        'page_obj': page_obj,
        'категории': категории,
        'производители': производители,
        'selected_category': категория_id,
        'selected_manufacturer': производитель_id,
        'search_query': поиск,
    }

    return render(request, 'shop/catalog.html', context)


def product_detail(request, pk):
    """
    Детальная информация о товаре по его ID
    """
    товар = get_object_or_404(
        Товар.objects.select_related('категория', 'производитель'),
        pk=pk
    )
    
    context = {
        'товар': товар,
    }
    
    return render(request, 'shop/product_detail.html', context)


@login_required
def add_to_cart(request, product_id):
    """
    Добавление товара в корзину
    Доступно только аутентифицированным пользователям
    """
    товар = get_object_or_404(Товар, pk=product_id)
    
    # Получаем или создаём корзину пользователя
    корзина, _ = Корзина.objects.get_or_create(пользователь=request.user)
    
    # Получаем количество из POST-запроса (по умолчанию 1)
    количество = int(request.POST.get('quantity', 1))
    
    # Проверяем, есть ли уже такой товар в корзине
    элемент, created = ЭлементКорзины.objects.get_or_create(
        корзина=корзина,
        товар=товар,
        defaults={'количество': количество}
    )
    
    if not created:
        # Если товар уже есть, увеличиваем количество
        элемент.количество += количество
        элемент.save()
    
    # Проверяем валидацию
    if элемент.количество > товар.количество_на_складе:
        messages.error(
            request,
            f'Недостаточно товара на складе. Доступно: {товар.количество_на_складе}'
        )
        элемент.количество = min(элемент.количество, товар.количество_на_складе)
        элемент.save()
    else:
        messages.success(request, f'Товар "{товар.название}" добавлен в корзину')
    
    return redirect('cart')


@login_required
def update_cart(request, item_id):
    """
    Обновление количества товара в корзине
    """
    элемент = get_object_or_404(ЭлементКорзины, pk=item_id)
    
    # Проверяем, что корзина принадлежит текущему пользователю
    if элемент.корзина.пользователь != request.user:
        messages.error(request, 'Вы не можете редактировать эту корзину')
        return redirect('cart')
    
    if request.method == 'POST':
        количество = int(request.POST.get('quantity', 1))
        
        # Валидация: количество не должно превышать наличие на складе
        if количество > элемент.товар.количество_на_складе:
            messages.error(
                request,
                f'Недостаточно товара на складе. Доступно: {элемент.товар.количество_на_складе}'
            )
            количество = элемент.товар.количество_на_складе
        
        if количество <= 0:
            # Если количество 0 или меньше, удаляем элемент
            элемент.delete()
            messages.info(request, 'Товар удалён из корзины')
        else:
            элемент.количество = количество
            элемент.save()
            messages.success(request, 'Количество товара обновлено')
    
    return redirect('cart')


@login_required
def remove_from_cart(request, item_id):
    """
    Удаление товара из корзины
    """
    элемент = get_object_or_404(ЭлементКорзины, pk=item_id)
    
    # Проверяем, что корзина принадлежит текущему пользователю
    if элемент.корзина.пользователь != request.user:
        messages.error(request, 'Вы не можете редактировать эту корзину')
        return redirect('cart')
    
    название_товара = элемент.товар.название
    элемент.delete()
    messages.success(request, f'Товар "{название_товара}" удалён из корзины')
    
    return redirect('cart')


@login_required
def cart_view(request):
    """
    Просмотр корзины пользователя
    """
    # Получаем корзину пользователя или создаём новую
    корзина, created = Корзина.objects.get_or_create(
        пользователь=request.user
    )
    
    # Получаем все элементы корзины
    элементы = корзина.элементы.select_related('товар').all()
    
    # Вычисляем общую стоимость
    общая_стоимость = корзина.общая_стоимость()
    
    context = {
        'корзина': корзина,
        'элементы': элементы,
        'общая_стоимость': общая_стоимость,
    }
    
    return render(request, 'shop/cart.html', context)


# ==================== Задание 3: Функционал оформления заказа ====================

def generate_order_excel(заказ):
    """
    Генерирует Excel файл с информацией о заказе
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ"
    
    # Стили
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws['A1'] = "ЧЕК ЗАКАЗА"
    ws['A1'].font = title_font
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Информация о заказе
    ws['A3'] = "Номер заказа:"
    ws['B3'] = f"#{заказ.id}"
    ws['A4'] = "Дата заказа:"
    ws['B4'] = заказ.дата_создания.strftime("%d.%m.%Y %H:%M")
    ws['A5'] = "Статус:"
    ws['B5'] = заказ.get_статус_display()
    
    # Информация о клиенте
    ws['A7'] = "ИНФОРМАЦИЯ О КЛИЕНТЕ"
    ws['A7'].font = Font(bold=True, size=11)
    ws['A8'] = "Имя:"
    ws['B8'] = заказ.пользователь.get_full_name() or заказ.пользователь.username
    ws['A9'] = "Email:"
    ws['B9'] = заказ.пользователь.email
    ws['A10'] = "Телефон:"
    ws['B10'] = заказ.телефон
    
    # Адрес доставки
    ws['A12'] = "АДРЕС ДОСТАВКИ"
    ws['A12'].font = Font(bold=True, size=11)
    ws['A13'] = "Адрес:"
    ws['B13'] = заказ.адрес_доставки
    ws['A14'] = "Город:"
    ws['B14'] = заказ.город
    ws['A15'] = "Почтовый индекс:"
    ws['B15'] = заказ.почтовый_индекс
    
    # Таблица товаров
    ws['A17'] = "ТОВАРЫ"
    ws['A17'].font = Font(bold=True, size=11)
    
    # Заголовок таблицы
    headers = ['Товар', 'Кол-во', 'Цена', 'Стоимость']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=18, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    # Элементы заказа
    row = 19
    for элемент in заказ.элементы.all():
        ws.cell(row=row, column=1).value = элемент.товар.название
        ws.cell(row=row, column=2).value = элемент.количество
        ws.cell(row=row, column=3).value = float(элемент.цена_на_момент_заказа)
        ws.cell(row=row, column=4).value = float(элемент.стоимость())
        
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border
            if col in [2, 3, 4]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='right')
        
        row += 1
    
    # Итоги
    total_row = row + 1
    ws.cell(row=total_row, column=1).value = "ИТОГО:"
    ws.cell(row=total_row, column=1).font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=4).value = float(заказ.сумма_заказа)
    ws.cell(row=total_row, column=4).font = Font(bold=True, size=11)
    for col in range(1, 5):
        ws.cell(row=total_row, column=col).border = border
    
    # Ширина колонок
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Сохраняем в BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def send_order_email(заказ):
    """
    Отправляет чек заказа по электронной почте
    """
    # Генерируем Excel файл
    excel_file = generate_order_excel(заказ)
    
    # Подготавливаем email
    subject = f"Ваш заказ #{заказ.id} в магазине настольных игр"
    message = f"""
    Спасибо за ваш заказ!
    
    Номер заказа: #{заказ.id}
    Дата заказа: {заказ.дата_создания.strftime("%d.%m.%Y %H:%M")}
    Сумма: {заказ.сумма_заказа} руб.
    
    Ваш чек прикреплён к этому письму.
    
    С уважением,
    Магазин настольных игр
    """
    
    from django.core.mail import EmailMessage
    
    email = EmailMessage(
        subject=subject,
        body=message,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[заказ.пользователь.email]
    )
    
    # Прикрепляем файл
    email.attach(
        f"check_order_{заказ.id}.xlsx",
        excel_file.read(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    try:
        email.send()
        return True
    except Exception as e:
        print(f"Ошибка при отправке email: {e}")
        return False


@login_required
def checkout(request):
    """
    Страница оформления заказа
    Доступна только аутентифицированным пользователям
    """
    # Получаем корзину пользователя
    try:
        корзина = Корзина.objects.get(пользователь=request.user)
    except Корзина.DoesNotExist:
        messages.error(request, 'Ваша корзина пуста')
        return redirect('cart')
    
    # Проверяем, есть ли товары в корзине
    элементы = корзина.элементы.select_related('товар').all()
    if not элементы.exists():
        messages.error(request, 'Ваша корзина пуста')
        return redirect('cart')
    
    if request.method == 'POST':
        # Получаем данные из формы
        адрес_доставки = request.POST.get('address', '').strip()
        город = request.POST.get('city', '').strip()
        почтовый_индекс = request.POST.get('postal_code', '').strip()
        телефон = request.POST.get('phone', '').strip()
        
        # Валидация данных
        if not all([адрес_доставки, город, почтовый_индекс, телефон]):
            messages.error(request, 'Пожалуйста, заполните все поля')
            context = {'корзина': корзина, 'элементы': элементы}
            return render(request, 'shop/checkout.html', context)
        
        # Создаём заказ
        сумма_заказа = корзина.общая_стоимость()
        заказ = Заказ.objects.create(
            пользователь=request.user,
            адрес_доставки=адрес_доставки,
            город=город,
            почтовый_индекс=почтовый_индекс,
            телефон=телефон,
            сумма_заказа=сумма_заказа,
            статус='pending'
        )
        
        # Копируем элементы корзины в заказ
        for элемент in элементы:
            ЭлементЗаказа.objects.create(
                заказ=заказ,
                товар=элемент.товар,
                количество=элемент.количество,
                цена_на_момент_заказа=элемент.товар.цена
            )
            
            # Уменьшаем количество товара на складе
            элемент.товар.количество_на_складе -= элемент.количество
            элемент.товар.save()
        
        # Отправляем email с чеком
        email_sent = send_order_email(заказ)
        
        # Очищаем корзину
        корзина.элементы.all().delete()
        
        messages.success(request, 'Заказ успешно оформлен!')
        if email_sent:
            messages.info(request, 'Чек отправлен на вашу почту')
        else:
            messages.warning(request, 'Чек не удалось отправить на почту')
        
        return redirect('order_success', order_id=заказ.id)
    
    # GET запрос - показываем форму
    общая_стоимость = корзина.общая_стоимость()
    
    context = {
        'корзина': корзина,
        'элементы': элементы,
        'общая_стоимость': общая_стоимость,
    }
    
    return render(request, 'shop/checkout.html', context)


@login_required
def order_success(request, order_id):
    """
    Страница успешного оформления заказа
    """
    заказ = get_object_or_404(Заказ, pk=order_id, пользователь=request.user)
    
    context = {
        'заказ': заказ,
    }
    
    return render(request, 'shop/order_success.html', context)


@login_required
def download_check(request, order_id):
    """
    Скачивание чека в формате Excel
    """
    заказ = get_object_or_404(Заказ, pk=order_id, пользователь=request.user)
    
    # Генерируем файл
    excel_file = generate_order_excel(заказ)
    
    # Возвращаем как attachment для скачивания
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="check_order_{заказ.id}.xlsx"'
    
    return response


@api_view(['GET', 'PATCH'])
@permission_classes([permissions.IsAuthenticated])
def me_api(request):
    profile = request.user.profile
    if request.method == 'GET':
        serializer = ProfileSerializer(profile)
        return Response(serializer.data)

    serializer = ProfileSerializer(profile, data=request.data, partial=True)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def cart_add_api(request):
    product_id = request.data.get('product_id')
    quantity = int(request.data.get('quantity', 1))

    if not product_id:
        return Response({'detail': 'Товар не указан'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        товар = Товар.objects.get(pk=product_id)
    except Товар.DoesNotExist:
        return Response({'detail': 'Товар не найден'}, status=status.HTTP_404_NOT_FOUND)

    if quantity <= 0:
        return Response({'detail': 'Количество должно быть больше нуля'}, status=status.HTTP_400_BAD_REQUEST)

    if товар.количество_на_складе < quantity:
        return Response({'detail': 'Недостаточно товара на складе'}, status=status.HTTP_400_BAD_REQUEST)

    корзина, _ = Корзина.objects.get_or_create(пользователь=request.user)
    элемент, created = ЭлементКорзины.objects.get_or_create(
        корзина=корзина,
        товар=товар,
        defaults={'количество': quantity}
    )
    if not created:
        элемент.количество = min(элемент.количество + quantity, товар.количество_на_складе)
        элемент.save()

    return Response(CartItemSerializer(элемент).data, status=status.HTTP_201_CREATED)


class ProductViewSet(viewsets.ModelViewSet):
    serializer_class = ProductSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        queryset = Товар.objects.select_related('категория', 'производитель').all()
        категория_id = self.request.query_params.get('category')
        производитель_id = self.request.query_params.get('manufacturer')
        поиск = self.request.query_params.get('search')

        if категория_id:
            queryset = queryset.filter(категория_id=категория_id)
        if производитель_id:
            queryset = queryset.filter(производитель_id=производитель_id)
        if поиск:
            queryset = queryset.filter(
                Q(название__icontains=поиск) | Q(описание__icontains=поиск)
            )
        return queryset


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Категория.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAdminOrReadOnly]


class ManufacturerViewSet(viewsets.ModelViewSet):
    queryset = Производитель.objects.all()
    serializer_class = ManufacturerSerializer
    permission_classes = [IsAdminOrReadOnly]


class CartViewSet(viewsets.ModelViewSet):
    serializer_class = CartSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Корзина.objects.filter(пользователь=self.request.user)

    def perform_create(self, serializer):
        serializer.save(пользователь=self.request.user)


class CartItemViewSet(viewsets.ModelViewSet):
    serializer_class = CartItemSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return ЭлементКорзины.objects.filter(корзина__пользователь=self.request.user)

    def perform_create(self, serializer):
        корзина, _ = Корзина.objects.get_or_create(пользователь=self.request.user)
        serializer.save(корзина=корзина)


class OrderViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        if self.request.user.is_staff:
            return Заказ.objects.select_related('пользователь').prefetch_related('элементы__товар').all()
        return Заказ.objects.filter(пользователь=self.request.user).prefetch_related('элементы__товар')
